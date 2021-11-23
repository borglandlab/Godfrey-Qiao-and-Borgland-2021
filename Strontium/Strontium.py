import os
from openpyxl import load_workbook, Workbook

os.chdir('/Users/nathansmac/Desktop/VTA_INPUTS/Strontium')
wb = load_workbook('Strontium.xlsx')

sheet_list = wb.sheetnames
sheet_list = sheet_list[0:-3]

ws_directory = wb.create_sheet('Directory')
ws_directory.cell(row=1, column=1).value = 'TabName'
ws_directory.cell(row=1, column=2).value = 'Trial'
ws_directory.cell(row=1, column=3).value = 'Blinded#'
ws_directory.cell(row=1, column=4).value = 'Length'
ws_directory.cell(row=1, column=5).value = 'Notes'


for n in range(len(sheet_list)):
	ws = wb[sheet_list[n]]
	ws_directory.cell(row=n+2, column=1).value = sheet_list[n]
	ws_directory.cell(row=n+2, column=2).value = ws.cell(row=1, column=1).value
	ws_directory.cell(row=n+2, column=3).value = ws.cell(row=1, column=2).value
	ws_directory.cell(row=n+2, column=4).value = ws.cell(row=3, column=1).value + 3
	ws_directory.cell(row=n+2, column=5).value = ws.cell(row=1, column=4).value

wb.save('Strontium.xlsx')
